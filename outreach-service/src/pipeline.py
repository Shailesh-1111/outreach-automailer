import pandas as pd
from openpyxl.styles import PatternFill
import os
import glob
import time
import random
import logging
from src.config import PROCESSING_QUEUE_DIR, PROCESSED_FILE
from src.template import generate_email_html
from src.sender import send_email

# Setup Email Logger
os.makedirs("logs", exist_ok=True)
email_logger = logging.getLogger("email_pipeline")
email_logger.setLevel(logging.INFO)
fh = logging.FileHandler("logs/email.log")
fh.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
if not email_logger.handlers:
    email_logger.addHandler(fh)

def get_latest_contacts_file():
    if not os.path.exists(PROCESSING_QUEUE_DIR):
        os.makedirs(PROCESSING_QUEUE_DIR, exist_ok=True)
        return None
    
    files = glob.glob(os.path.join(PROCESSING_QUEUE_DIR, "*.csv"))
    if not files:
        return None
        
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def run_outreach(target_file=None, mode='all', role="SDE", template_type="formal", profile=None):
    if profile is None: profile = {}
    sender_name = profile.get('name', 'Shailesh Yadav')
    sender_exp = profile.get('experience', '1.5+ years')
    sender_email = profile.get('email', 'shailesh112001y@gmail.com')
    if target_file is None:
        target_file = get_latest_contacts_file()
    if not target_file:
        print(f"❌ Error: No CSV files found in {PROCESSING_QUEUE_DIR}/ folder.")
        return

    df = pd.read_csv(target_file)
    if "verdict" not in df.columns:
        df["verdict"] = ""
    if "verdict_group" not in df.columns:
        df["verdict_group"] = ""
    if "is_sent" not in df.columns:
        df["is_sent"] = ""

    # LOAD GLOBAL HISTORY TO PREVENT DUPLICATES
    global_history_df = None
    sent_emails_globally = set()
    if os.path.exists(PROCESSED_FILE):
        try:
            global_history_df = pd.read_excel(PROCESSED_FILE)
            # Find everything marked Success globally
            if 'email' in global_history_df.columns and 'verdict_group' in global_history_df.columns:
                sent_mask = global_history_df['verdict_group'].astype(str).str.lower() == 'success'
                sent_emails_globally = set(global_history_df.loc[sent_mask, 'email'].dropna().astype(str).str.strip())
            # Fallback for old version
            elif 'email' in global_history_df.columns and 'verdict' in global_history_df.columns:
                sent_mask = global_history_df['verdict'].astype(str).str.lower() == 'sent'
                sent_emails_globally = set(global_history_df.loc[sent_mask, 'email'].dropna().astype(str).str.strip())
                
            print(f"🌍 Loaded {len(global_history_df)} historical records. {len(sent_emails_globally)} already sent globally.")
        except Exception as e:
            print(f"Warning: Could not read {PROCESSED_FILE}: {e}")

    processed_records = []
    print(f"📋 Loaded {len(df)} contacts from {target_file}.\n")

    for idx, row in df.iterrows():
        email_to = str(row.get("email", "")).strip()
        name = str(row.get("first_name", "")).strip()
        company = str(row.get("company", "")).strip()
        v_group_str = str(row.get("verdict_group", "")).strip().lower()
        v_str = str(row.get("verdict", "")).strip().lower()
        
        if pd.isna(email_to) or not email_to or str(email_to).lower() == 'nan':
            continue

        if not name or str(name).lower() == 'nan':
            print(f"⏩ Skipping {email_to} — Missing first name.")
            row["is_sent"] = False
            row["verdict_group"] = "Data Error"
            row["verdict"] = "Missing first name in CSV"
            processed_records.append(row)
            continue

        if not company or str(company).lower() == 'nan':
            print(f"⏩ Skipping {email_to} — Missing company name.")
            row["is_sent"] = False
            row["verdict_group"] = "Data Error"
            row["verdict"] = "Missing company name in CSV"
            processed_records.append(row)
            continue

        # Mode filters
        if mode == 'failed':
            # Skip if it is not an error
            if "error" not in v_group_str and "failed" not in v_group_str and "error" not in v_str and "failed" not in v_str:
                continue
                
        elif mode == 'unprocessed':
            # Skip if it has ANY group (it was processed in some way)
            if v_group_str != '' and v_group_str != 'pending' and v_str != '' and v_str != 'pending':
                continue
                
        elif mode != 'all':
            # Exact match for custom groups like 'env error', 'auth error'
            if mode != v_group_str and mode != v_str:
                continue

        try:
            # Skip if already marked sent LOCALLY
            if str(row.get("is_sent", "")).lower() == "true":
                email_logger.info(f"[SKIPPED LOCAL] {email_to}")
                print(f"⏩ [Skipped] {email_to}")
                row["verdict_group"] = "Skipped"
                row["verdict"] = "Already dispatched successfully in a previous run."
                processed_records.append(row)
                continue
                
            # Skip if already marked sent GLOBALLY
            if email_to in sent_emails_globally:
                email_logger.info(f"[SKIPPED GLOBAL] {email_to}")
                print(f"🌍 [Skipped] {email_to}")
                row["is_sent"] = True
                row["verdict_group"] = "Skipped (Global)"
                row["verdict"] = "Found in global history. Prevented duplicate email."
                processed_records.append(row)
                continue
    
            # Generate and send email
            email_body = generate_email_html(name, company, role=role, template_type=template_type, sender_name=sender_name, sender_exp=sender_exp, sender_email=sender_email)
            subject = f"{sender_name} | Exploring {role} Opportunities at {company}"
            
            success, v_group, v_msg = send_email(email_to, subject, email_body)
    
            row["is_sent"] = success
            row["verdict_group"] = v_group
            row["verdict"] = v_msg
    
            processed_records.append(row)
            
            if success:
                email_logger.info(f"[SUCCESS] {email_to}")
                print(f"✅ [Sent] {email_to}")
                delay = random.uniform(2, 5)
                time.sleep(delay)
            else:
                email_logger.error(f"[{v_group.upper()}] {email_to} - {v_msg}")
                print(f"❌ [{v_group}] {email_to}")
                # If a critical error happens, halt the entire job immediately
                if v_group in ["Env Error", "Auth Error", "System Error"]:
                    print(f"🛑 Critical Error ({v_group}): Aborting remaining pipeline to prevent cascading failures.")
                    break
            
        except Exception as e:
            print(f"❌ Crash intercepted for {email_to}: {e}")
            row["is_sent"] = False
            row["verdict_group"] = "System Error"
            row["verdict"] = f"Pipeline Crash: {str(e)}"
            processed_records.append(row)
            print("🛑 Pipeline crashed. Aborting remaining rows.")
            break

    if not processed_records:
        print("No new records to process.")
        return

    # Create DataFrames
    processed_df = pd.DataFrame(processed_records)
    
    # Save unprocessed back to CSV
    for i, p_row in processed_df.iterrows():
        email = p_row['email']
        idx = df[df['email'] == email].index
        if not idx.empty:
            df.loc[idx[0], 'is_sent'] = p_row['is_sent']
            df.loc[idx[0], 'verdict_group'] = p_row['verdict_group']
            df.loc[idx[0], 'verdict'] = p_row['verdict']

    df.to_csv(target_file, index=False)
    print(f"\n📤 Contacts updated and saved to {target_file}")

    # Combine with global history
    if global_history_df is not None and not global_history_df.empty:
        # Prevent global skips from overwriting original success
        new_records_for_excel = processed_df[processed_df['verdict_group'] != 'Skipped (Global)']
        combined_df = pd.concat([global_history_df, new_records_for_excel], ignore_index=True)
        combined_df = combined_df.drop_duplicates(subset=['email'], keep='last')
    else:
        combined_df = processed_df

    # Ensure the directory exists
    os.makedirs(os.path.dirname(PROCESSED_FILE), exist_ok=True)

    # Save processed to Excel with color verdict
    with pd.ExcelWriter(PROCESSED_FILE, engine="openpyxl") as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Processed")
        workbook = writer.book
        sheet = writer.sheets["Processed"]

        try:
            # Color by verdict_group if it exists, else verdict
            target_col = "verdict_group" if "verdict_group" in combined_df.columns else "verdict"
            verdict_col = combined_df.columns.get_loc(target_col) + 1
            
            for i, group in enumerate(combined_df[target_col], start=2):
                group_val = str(group).lower()
                if "success" in group_val or group_val == "sent":
                    color = "C6EFCE"  # green
                elif "error" in group_val or "failed" in group_val:
                    color = "FFC7CE"  # red
                elif "skipped" in group_val:
                    color = "FFEB9C"  # yellow
                else:
                    color = "FFFFFF"
                cell = sheet.cell(row=i, column=verdict_col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        except KeyError:
            pass

    print(f"✅ Processed records Excel report saved to {PROCESSED_FILE}")
